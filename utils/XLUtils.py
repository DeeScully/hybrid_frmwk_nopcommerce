import openpyxl

def get_row_count(file, sheetname):
    wkbk = openpyxl.load_workbook(file)
    sheet = wkbk[sheetname]
    return sheet.max_row

def get_col_count(file, sheetname):
    wkbk = openpyxl.load_workbook(file)
    sheet = wkbk[sheetname]
    return sheet.max_column

def read_data(file, sheetname, rownum, colnum):
    wkbk = openpyxl.load_workbook(file)
    sheet = wkbk[sheetname]
    return sheet.cell(row=rownum, column=colnum).value

def write_data(file, sheetname, rownum, colnum, data):
    wkbk = openpyxl.load_workbook(file)
    sheet = wkbk[sheetname]
    sheet.cell(row=rownum, column=colnum).value = data
    wkbk.save(file)